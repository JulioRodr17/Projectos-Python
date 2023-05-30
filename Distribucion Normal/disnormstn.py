import math
import openpyxl
from scipy import integrate

# lambda es igual a x
# miu es igual a m
def integral(y):
    return (1 / (math.sqrt(2 * math.pi))) * math.exp(-0.5 * (y ** 2))

def calcula_integral(b):
    resultado = integrate.quad(integral, -6, b)[0]
    return resultado

hd= openpyxl.load_workbook('pruebas.xlsx')
hoja = hd.active
j,k=1 ,2
b= -5.98
while j<=30 and b<=6.00:
    if(hoja.cell(row=1, column=j)) and j%2!=0 :
        hoja.cell(row=1, column=j).value = 'Indice'
        hoja.cell(row=1, column=j+1).value = 'Numeros'

    if(hoja.cell(row=k, column=j) and j%2!=0):
        integral_resultado = calcula_integral(b)
        hoja.cell(row=k, column=j).value=b
        hoja.cell(row=k, column=j+1).value=integral_resultado
        b+=0.02
    elif b==0:
        hoja.cell(row=k, column=j).value=0
    k+=1
    if(k/48==1):
        j+=1
        k=2

hd.save('pruebas.xlsx')

